#!/usr/bin/env python3
"""Render frontier-tracker scan/state data into user-selected display forms."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import html
import json
import re
import sys
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_STATE = ROOT / "state" / "reading_state.json"
DEFAULT_OUTPUTS = ROOT / "outputs"
DISPLAY_MODES = ("excel", "app", "notes", "codex")
CORE_TERMS = (
    "ecosystem diversity",
    "ecosystem type",
    "ecosystem types",
    "ecosystem classification",
    "ecosystem typology",
    "ecosystem unit",
    "ecosystem units",
    "ecosystem mapping",
    "ecosystem map",
)
BOUNDARY_TERMS = (
    "landscape diversity",
    "landscape heterogeneity",
    "land-use diversity",
    "land use diversity",
    "land cover diversity",
    "habitat diversity",
    "habitat heterogeneity",
    "spectral diversity",
    "spectral heterogeneity",
    "proxy indicator",
    "surrogate indicator",
)
METHOD_TERMS = (
    "remote sensing",
    "satellite",
    "imagery",
    "earth observation",
    "land cover",
    "land use",
    "land-use",
    "spatial",
    "mapping",
    "gis",
    "dataset",
    "data set",
    "classification",
)
ECO_ENV_TERMS = (
    "biodiversity",
    "conservation",
    "ecosystem service",
    "ecosystem services",
    "vegetation",
    "forest",
    "grassland",
    "wetland",
    "soil",
    "carbon",
    "biomass",
    "npp",
    "gpp",
    "water quality",
    "watershed",
    "river",
    "climate",
    "sustainability",
    "urban",
    "agriculture",
    "food system",
)
NOISE_TERMS = (
    "cancer",
    "tumour",
    "tumor",
    "car-t",
    "leukemia",
    "lymphoma",
    "metastatic",
    "prostate",
    "bladder pain",
    "placental",
    "mitochondrial",
    "crispr",
    "clinical trial",
    "antibody",
    "vaccine",
    "immun",
    "single-cell",
    "protein",
    "rna",
    "dna",
    "graphene",
    "perovskite",
    "photovoltaic",
    "hydrogel",
    "quantum",
)


def today_stamp() -> str:
    return dt.date.today().isoformat()


def slugify(text: str, fallback: str = "paper") -> str:
    value = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff]+", "-", text).strip("-").lower()
    return value[:90] or fallback


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def paper_key(record: dict[str, Any], index: int) -> str:
    return record.get("key") or record.get("doi") or record.get("openalex_id") or f"paper-{index + 1}"


def contains_any(text: str, terms: tuple[str, ...]) -> bool:
    return any(term in text for term in terms)


def classify_for_profile(record: dict[str, Any]) -> dict[str, str]:
    text = " ".join(
        str(record.get(key, "")) for key in ("title", "journal", "topics", "relevance", "family", "pool")
    ).lower()
    if contains_any(text, CORE_TERMS):
        if "classification" in text or "typology" in text or "unit" in text:
            klass = "B"
        elif "mapping" in text or "map" in text or "spatial" in text:
            klass = "D"
        else:
            klass = "A"
        return {
            "relevance": "high",
            "priority": "P1",
            "class": klass,
            "status": "queued",
            "action": "read-note",
        }
    if contains_any(text, BOUNDARY_TERMS):
        return {
            "relevance": "medium",
            "priority": "P2",
            "class": "F",
            "status": "to-screen",
            "action": "boundary-check",
        }
    method_hit = contains_any(text, METHOD_TERMS)
    env_hit = contains_any(text, ECO_ENV_TERMS)
    if method_hit and env_hit:
        return {
            "relevance": "medium",
            "priority": "P2",
            "class": "C" if "dataset" in text or "classification" in text else "D",
            "status": "to-screen",
            "action": "screen-read",
        }
    if env_hit:
        return {
            "relevance": "medium",
            "priority": "P3",
            "class": "E",
            "status": "to-screen",
            "action": "screen",
        }
    if contains_any(text, NOISE_TERMS):
        return {
            "relevance": "low",
            "priority": "P3",
            "class": "noise",
            "status": "skip",
            "action": "skip",
        }
    return {
        "relevance": "low",
        "priority": "P3",
        "class": "noise",
        "status": "skip",
        "action": "skip",
    }


def records_from_scan(scan: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for bucket in ("new", "already_seen", "needs_manual_verification"):
        for record in scan.get(bucket) or []:
            item = dict(record)
            item.setdefault("scan_bucket", bucket)
            records.append(item)
    return records


def records_from_state(state: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    queue = state.get("reading_queue") or {}
    for index, (key, record) in enumerate((state.get("seen") or {}).items()):
        item = dict(record)
        item.setdefault("key", key)
        item.setdefault("scan_bucket", "state")
        queue_record = queue.get(key) or {}
        if queue_record.get("status"):
            item["status"] = queue_record["status"]
        records.append(item)
    return records


def load_records(scan_json: Path | None, state: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if scan_json:
        scan = load_json(scan_json)
        return records_from_scan(scan), {"source": str(scan_json), "scan": scan}
    state_data = load_json(state)
    return records_from_state(state_data), {"source": str(state), "state": state_data}


def normalize_record(record: dict[str, Any], index: int) -> dict[str, Any]:
    classification = classify_for_profile(record)
    return {
        "key": paper_key(record, index),
        "title": record.get("title", ""),
        "journal": record.get("journal", ""),
        "publication_date": record.get("publication_date", ""),
        "authors": record.get("authors", ""),
        "doi": record.get("doi", ""),
        "doi_url": record.get("doi_url", ""),
        "openalex_id": record.get("openalex_id", ""),
        "pool": record.get("pool", ""),
        "family": record.get("family", ""),
        "pool_source": record.get("pool_source", ""),
        "rank_basis": record.get("rank_basis", ""),
        "topics": record.get("topics", ""),
        "relevance": record.get("relevance") if record.get("relevance") not in {"", None, "needs-profile-screening"} else classification["relevance"],
        "priority": record.get("P1_P3_priority") or record.get("priority") or classification["priority"],
        "class": record.get("A_F_class") or record.get("class") or classification["class"],
        "status": record.get("status") if record.get("status") not in {"", None, "to-screen"} else classification["status"],
        "action": record.get("action") or classification["action"],
        "scan_bucket": record.get("scan_bucket", ""),
    }


def normalized_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [normalize_record(record, index) for index, record in enumerate(records)]


def ensure_output(path: Path | None, default: Path) -> Path:
    target = path or default
    target.parent.mkdir(parents=True, exist_ok=True)
    return target


def render_excel(records: list[dict[str, Any]], output: Path | None) -> Path:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    target = ensure_output(output, DEFAULT_OUTPUTS / "excel" / f"frontier_dashboard_{today_stamp()}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "01_Dashboard"

    headers = [
        "priority",
        "class",
        "status",
        "action",
        "publication_date",
        "journal",
        "title",
        "relevance",
        "pool",
        "topics",
        "doi_url",
        "rank_basis",
        "scan_bucket",
    ]
    rows = normalized_records(records)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    widths = {
        "A": 12,
        "B": 10,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 26,
        "G": 70,
        "H": 24,
        "I": 28,
        "J": 48,
        "K": 40,
        "L": 32,
        "M": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        row[10].hyperlink = row[10].value if row[10].value else None
        if row[10].hyperlink:
            row[10].style = "Hyperlink"

    red = PatternFill("solid", fgColor="F4CCCC")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    green = PatternFill("solid", fgColor="D9EAD3")
    gray = PatternFill("solid", fgColor="E7E6E6")
    orange = PatternFill("solid", fgColor="FCE4D6")
    max_row = max(2, ws.max_row)
    ws.conditional_formatting.add(f"A2:A{max_row}", FormulaRule(formula=['$A2="P1"'], fill=red))
    ws.conditional_formatting.add(f"A2:A{max_row}", FormulaRule(formula=['$A2="P2"'], fill=yellow))
    ws.conditional_formatting.add(f"H2:H{max_row}", FormulaRule(formula=['$H2="high"'], fill=green))
    ws.conditional_formatting.add(f"H2:H{max_row}", FormulaRule(formula=['$H2="needs-profile-screening"'], fill=orange))
    ws.conditional_formatting.add(f"C2:C{max_row}", FormulaRule(formula=['OR($C2="skip",$C2="archived")'], fill=gray))

    for title, predicate in [
        ("02_New", lambda r: r.get("scan_bucket") == "new"),
        ("03_Queue", lambda r: r.get("status") in {"to-screen", "queued", "reading"}),
        ("04_Boundary_Proxy", lambda r: any(x in (r.get("topics", "") + " " + r.get("relevance", "")).lower() for x in ["proxy", "boundary", "landscape", "habitat", "spectral"])),
    ]:
        sheet = wb.create_sheet(title)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="548235")
        for row in [r for r in rows if predicate(r)]:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for idx, width in enumerate([12, 10, 14, 14, 14, 26, 70, 24, 28, 48, 40, 32, 16], 1):
            sheet.column_dimensions[get_column_letter(idx)].width = width

    wb.save(target)
    return target


def render_app(records: list[dict[str, Any]], output: Path | None) -> Path:
    app_dir = output or (DEFAULT_OUTPUTS / "app" / f"frontier_table_app_{today_stamp()}")
    app_dir.mkdir(parents=True, exist_ok=True)
    rows = normalized_records(records)
    (app_dir / "data.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    index = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Frontier Tracker</title>
  <style>
    body { font-family: system-ui, -apple-system, Segoe UI, sans-serif; margin: 0; color: #202124; background: #f7f8fa; }
    header { padding: 18px 24px; background: #1f4e78; color: white; }
    main { padding: 18px 24px; }
    .toolbar { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 14px; }
    input, select { padding: 8px 10px; border: 1px solid #c9d1d9; border-radius: 6px; background: white; }
    table { width: 100%; border-collapse: collapse; background: white; font-size: 13px; }
    th, td { padding: 8px 10px; border-bottom: 1px solid #e5e7eb; vertical-align: top; }
    th { position: sticky; top: 0; background: #eef3f8; text-align: left; }
    tr:hover { background: #f4f8fb; }
    .badge { display: inline-block; padding: 2px 7px; border-radius: 999px; background: #e8f0fe; }
    .title { min-width: 360px; }
    a { color: #0b57d0; }
  </style>
</head>
<body>
  <header><h1>Frontier Tracker</h1></header>
  <main>
    <div class="toolbar">
      <input id="q" placeholder="搜索标题、期刊、主题" size="34">
      <select id="status"><option value="">全部状态</option></select>
      <select id="pool"><option value="">全部池</option></select>
      <select id="bucket"><option value="">全部来源</option></select>
    </div>
    <div id="count"></div>
    <table>
      <thead><tr><th>日期</th><th>期刊</th><th class="title">标题</th><th>状态</th><th>优先级</th><th>池</th><th>主题</th><th>DOI</th></tr></thead>
      <tbody id="rows"></tbody>
    </table>
  </main>
  <script>
    let papers = [];
    const fields = ["status", "pool", "scan_bucket"];
    function optionize(id, values) {
      const el = document.getElementById(id);
      [...new Set(values.filter(Boolean))].sort().forEach(v => {
        const opt = document.createElement("option");
        opt.value = v; opt.textContent = v; el.appendChild(opt);
      });
    }
    function render() {
      const q = document.getElementById("q").value.toLowerCase();
      const status = document.getElementById("status").value;
      const pool = document.getElementById("pool").value;
      const bucket = document.getElementById("bucket").value;
      const filtered = papers.filter(p => {
        const hay = [p.title, p.journal, p.topics, p.authors].join(" ").toLowerCase();
        return (!q || hay.includes(q)) && (!status || p.status === status) && (!pool || p.pool === pool) && (!bucket || p.scan_bucket === bucket);
      });
      document.getElementById("count").textContent = `显示 ${filtered.length} / ${papers.length} 篇`;
      document.getElementById("rows").innerHTML = filtered.map(p => `
        <tr>
          <td>${p.publication_date || ""}</td>
          <td>${p.journal || ""}</td>
          <td class="title">${p.title || ""}</td>
          <td><span class="badge">${p.status || ""}</span></td>
          <td>${p.priority || ""}</td>
          <td>${p.pool || ""}</td>
          <td>${p.topics || ""}</td>
          <td>${p.doi_url ? `<a href="${p.doi_url}" target="_blank">DOI</a>` : ""}</td>
        </tr>`).join("");
    }
    fetch("data.json").then(r => r.json()).then(data => {
      papers = data;
      optionize("status", papers.map(p => p.status));
      optionize("pool", papers.map(p => p.pool));
      optionize("bucket", papers.map(p => p.scan_bucket));
      document.querySelectorAll("input,select").forEach(el => el.addEventListener("input", render));
      render();
    });
  </script>
</body>
</html>
"""
    (app_dir / "index.html").write_text(index, encoding="utf-8")
    return app_dir


def note_text(record: dict[str, Any]) -> str:
    title = record.get("title", "Untitled")
    return f"""---
title: "{title.replace('"', "'")}"
journal: "{record.get('journal', '')}"
published: "{record.get('publication_date', '')}"
doi: "{record.get('doi', '')}"
status: "{record.get('status', 'to-screen')}"
priority: "{record.get('priority', '')}"
class: "{record.get('class', '')}"
relevance: "{record.get('relevance', '')}"
pool: "{record.get('pool', '')}"
---

# {title}

## 一句话判断
- 待筛选。

## 基本信息
- 期刊：{record.get('journal', '')}
- 日期：{record.get('publication_date', '')}
- DOI：{record.get('doi_url', '')}
- 作者：{record.get('authors', '')}

## 与我研究的关系
- 直接关系：
- 间接关系：
- 县域尺度可用性：
- 是否涉及生态系统多样性本体：

## 可提取内容
- 概念：
- 方法：
- 指标：
- 数据：
- 讨论观点：

## 边界检查
- 是否把 land-use type 当 ecosystem type：
- 是否把 landscape diversity 当 ecosystem diversity：
- 是否只是 proxy：

## 我的下一步
- read / cite / note / archive / skip
"""


def render_notes(records: list[dict[str, Any]], output: Path | None, obsidian_dir: Path | None, zotero_dir: Path | None) -> Path:
    base = output or (DEFAULT_OUTPUTS / "notes" / today_stamp())
    obsidian = obsidian_dir or (base / "obsidian_notes")
    zotero = zotero_dir or (base / "zotero_notes")
    obsidian.mkdir(parents=True, exist_ok=True)
    zotero.mkdir(parents=True, exist_ok=True)
    rows = normalized_records(records)
    for index, record in enumerate(rows):
        name = f"{record.get('publication_date') or today_stamp()}_{slugify(record.get('title', ''), f'paper-{index+1}')}.md"
        text = note_text(record)
        (obsidian / name).write_text(text, encoding="utf-8")
        (zotero / name).write_text(text, encoding="utf-8")
    csv_path = zotero / "zotero_note_index.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as fh:
        fieldnames = ["title", "journal", "publication_date", "doi", "doi_url", "status", "priority", "class", "relevance"]
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for record in rows:
            writer.writerow({key: record.get(key, "") for key in fieldnames})
    return base


def render_codex(records: list[dict[str, Any]], output: Path | None, limit: int) -> Path:
    target = ensure_output(output, DEFAULT_OUTPUTS / "codex" / f"frontier_preview_{today_stamp()}.md")
    rows = normalized_records(records)
    lines = [
        "# Frontier Tracker Preview",
        f"Generated: {today_stamp()}",
        f"Total papers: {len(rows)}",
        "",
        "## Top Items",
        "| Date | Journal | Title | Status | Pool | DOI |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for record in rows[:limit]:
        doi = record.get("doi_url", "")
        lines.append(f"| {record.get('publication_date', '')} | {record.get('journal', '')} | {record.get('title', '')} | {record.get('status', '')} | {record.get('pool', '')} | {doi} |")
    lines.extend(["", "## Action List", "- Read:", "- Note:", "- Archive or skip:"])
    target.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return target


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Render frontier-tracker data into the user-selected display form.")
    parser.add_argument("--mode", choices=DISPLAY_MODES, required=True)
    parser.add_argument("--scan-json", type=Path, help="Scan JSON from scan_recent_papers.py.")
    parser.add_argument("--state", type=Path, default=DEFAULT_STATE, help="State file used when --scan-json is omitted.")
    parser.add_argument("--output", type=Path, help="Output file or directory, depending on mode.")
    parser.add_argument("--obsidian-dir", type=Path, help="Destination folder for Obsidian notes in notes mode.")
    parser.add_argument("--zotero-dir", type=Path, help="Destination folder for Zotero-note-ready exports in notes mode.")
    parser.add_argument("--limit", type=int, default=25, help="Maximum records in Codex preview.")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    records, meta = load_records(args.scan_json, args.state)
    if args.mode == "excel":
        result = render_excel(records, args.output)
    elif args.mode == "app":
        result = render_app(records, args.output)
    elif args.mode == "notes":
        result = render_notes(records, args.output, args.obsidian_dir, args.zotero_dir)
    else:
        result = render_codex(records, args.output, args.limit)
    sys.stdout.write(json.dumps({"mode": args.mode, "records": len(records), "source": meta["source"], "output": str(result)}, ensure_ascii=False, indent=2) + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
